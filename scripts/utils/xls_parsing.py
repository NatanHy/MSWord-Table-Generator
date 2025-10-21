from functools import cache
from typing import List, Dict, Iterable

import openpyxl
import pandas as pd

from table_generation import Component
from utils.dataframes import make_first_row_headers
from utils.caching import cache_on_attr
from utils.files import ExcelFileManager

@cache
def parse_excel_cached(xls_path : str) -> ExcelFileManager:
    return ExcelFileManager(xls_path)

def get_description(file_manager : ExcelFileManager, component_id : str) -> str:
    ws = file_manager.wb[component_id]
    return ws["C14"].value

def set_description(file_manager : ExcelFileManager, component_id : str, description : str):
    file_manager.write(component_id, "C14", description)

def get_filtered_by_id(file_manager : ExcelFileManager, prefix="") -> pd.DataFrame:
    # Get main sheet
    xls = file_manager.xls
    try:
        df = xls.parse("PSAR SFK FEP list", header=None)
    except:
        df = xls.parse("SFK FEP list", header=None)

    # Skip to row where SKB FEP ID is located
    col_b = df.columns[1]
    offset = df[df[col_b] == "SKB FEP ID"].index[0]
    df_skipped = df[offset:]

    # Filter SKB FEP ID, FEP Name and System Component columns
    df_filtered = make_first_row_headers(df_skipped)[["SKB FEP ID", "FEP Name", "System Component", "Description"]]
    var_prefix = df_filtered["SKB FEP ID"].dropna().iloc[0] # Prefix like Ge, Bio, C etc.

    # Filter variables like Ge01 or Bio01 etc.
    filtered_by_id = df_filtered[df_filtered["SKB FEP ID"].str.match(rf"{prefix}{var_prefix}[0-9]+", na=False)]
    return filtered_by_id

def get_component_by_id(file_manager : ExcelFileManager, id : str) -> Component:
    filtered_by_id = get_filtered_by_id(file_manager)
    row = filtered_by_id[filtered_by_id.iloc[:, 0] == id]

    try:
        c_id = row["SKB FEP ID"].iloc[0]
        name = row["FEP Name"].iloc[0]
        system_component = row["System Component"].iloc[0]
        return Component(file_manager, c_id, name, system_component)
    except:
        raise ValueError("Invalid component ID")

def parse_components(file_manager : ExcelFileManager) -> List[Component]:
    """
    Parse Component info from the PSAR SKF FEP list sheet of the excel file. 

    ### Parameters
    xls : `pd.ExcelFile` specifying Component data in the PSAR SKF FEP list sheet

    ### Returns
    List of `Component` objects. 
    """

    filtered_by_id = get_filtered_by_id(file_manager)
    components = []

    for _, row in filtered_by_id.iterrows():
        id = row["SKB FEP ID"]
        name = row["FEP Name"]
        system_component = row["System Component"]
        components.append(Component(file_manager, id, name, system_component))

    return components

def parse_variables(file_manager : ExcelFileManager) -> Dict[str, str]:
    """
    Parse Variables from the PSAR SKF FEP list sheet of the excel file. 

    ### Parameters
    xls : `pd.ExcelFile` specifying Variable data in the PSAR SKF FEP list sheet

    ### Returns
    Dictionary from variables to Component names. 
    """

    filtered_by_id = get_filtered_by_id(file_manager, prefix="Var")
    variables = {}

    for _, row in filtered_by_id.iterrows():
        id = row["SKB FEP ID"]
        name = row["FEP Name"]
        variables[id] = name

    return variables

@cache
def _xls_matches_prefix(prefix: str, xls_file_path : str) -> bool:
    wb = openpyxl.load_workbook(xls_file_path, data_only=True, read_only=True)
    try:
        ws = wb["PSAR SFK FEP list"]
    except:
        ws = wb["SFK FEP list"]
    return ws["B8"].value == prefix

def get_xls_from_component_id(component_id : str, xls_files : Iterable[str]) -> str | None: 
    process_prefix = component_id[0]

    for pth in xls_files:
        if _xls_matches_prefix(process_prefix, pth):
            return pth